"""
Circuit Schedule File Parser — Gemini AI 기반 해석.

다양한 형식의 회로 스케줄 파일을 Gemini AI로 해석하여 구조화된 JSON으로 변환.

지원 형식:
  - Excel (.xlsx, .xls): openpyxl로 셀 데이터 → 텍스트 변환 → Gemini 텍스트 해석
  - CSV (.csv): 텍스트로 읽기 → Gemini 텍스트 해석
  - Image (.jpg, .png): Gemini Vision API 직접 해석
  - PDF (.pdf): Gemini Vision API 직접 해석
"""

from __future__ import annotations

import base64
import csv
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)


# ── Gemini Prompt ─────────────────────────────────────

SCHEDULE_EXTRACTION_PROMPT = """You are an expert electrical engineer specializing in Singapore electrical installations and Single Line Diagrams (SLD).

Your task: Analyze the uploaded circuit schedule file (Excel spreadsheet, CSV table, or image/photo of a circuit schedule) and extract ALL electrical information into structured JSON.

## What to Extract

### 1. Main Circuit / Incoming Supply Information
Look for these fields (may appear as key-value pairs or in a header section):
- Supply source (e.g., "SP PowerGrid", "Landlord Riser", "From Landlord Supply")
- Supply type (single phase 230V or three phase 400V)
- Approved load / kVA rating
- Main breaker (type, rating, poles, fault rating kA, trip characteristic)
- ELCB/RCCB (type, rating, sensitivity mA, poles)
- Incoming cable specification
- Busbar rating
- Metering type (CT meter, SP meter, KWH meter)
- CT ratio (if applicable)

### 2. Multiple Distribution Boards (Multi-DB)
If the file has multiple sheets/tabs or sections representing separate distribution boards (e.g., MSB + DB2, Lighting DB + Power DB):
- Each board has its own breaker, ELCB, busbar, and sub-circuits
- Use `distribution_boards` array instead of `outgoing_circuits`
- Look for "Protection Group" columns — these indicate per-phase RCCB grouping within a DB

### 3. Circuit Schedule / Sub-circuits Table
Extract EVERY row from the circuit table:
- Circuit ID/name (e.g., L1P1, S1, P3, ISOL 1, Spare)
- Load description (e.g., "8 Nos Lighting Points", "6 Nos 13A Twin S/S/O", "1 no. 20A DP isolator")
- Breaker specification (type, rating, poles, fault kA, characteristic)
- Cable specification (size, type, cores, CPC, installation method)
- Room/location (if specified)
- Phase assignment (L1/L2/L3 or R/Y/B)
- Protection group (if present — e.g., "RCCB L1", "RCCB L2", "RCCB L3")

### 4. Client / Project Information (if present)
- Client name, address, unit number
- Drawing number, contractor info

## Singapore Electrical Standards Reference
- Single-phase: 230V, DP/SPN poles
- Three-phase: 400V, TPN/4P poles
- MCB: ≤ 100A, MCCB: 125–630A, ACB: > 630A
- Sub-circuit MCB fault rating: typically 6kA
- Main breaker MCB fault rating: typically 10kA
- ELCB/RCCB: mandatory per SS 638
- Common cable: "2 x 1C 2.5mm² PVC + 2.5mm² CPC in G.I. conduit / metal trunking"

## Supply Source & Metering Rules
- supply_source and metering are INDEPENDENT properties:
  - supply_source: WHERE the power comes from → determines incoming_label, isolator type
  - metering: HOW metering is done → determines CT metering section rendering
- supply_source: "sp_powergrid" (SP PowerGrid direct supply) or "landlord" (building riser / landlord supply)
- **IMPORTANT — supply_source and metering are determined independently**:
  - supply_source is determined ONLY from supply text keywords (building riser → "landlord", SP/HDB → "sp_powergrid"). Do NOT change supply_source based on metering equipment.
  - metering is determined from explicit metering equipment in the schedule:
    - "SPPG kWh meter" with CT ratio (e.g. "100/5A CT") → metering type = "ct_meter"
    - "SPPG kWh meter" without CT → metering type = "sp_meter"
    - No metering equipment mentioned → metering = null
  - A landlord supply CAN have CT metering (e.g., SPPG kWh + CT in a building riser installation). This is common for larger commercial tenants.
- **Landlord supply**: Keywords: "SUPPLY FROM BUILDING RISER", "LANDLORD", "FROM RISER"
  - Landlord installations have their own isolator inside the unit (enclosed type), NOT an SP meter board.
- **SP PowerGrid supply** (direct SP supply): metering is "sp_meter" (residential) or "ct_meter" (≥125A three-phase).
  - Keywords: "INCOMING FROM HDB", "SP POWERGRID", "FROM SP"
- If supply source is unclear, default to "sp_powergrid".
- **incoming_label**: Extract from supply source text for the SLD diagram label. Common patterns:
  - "Landlord riser ..." → "FROM LANDLORD RISER"
  - "HDB electrical riser ..." → "INCOMING FROM HDB ELECTRICAL RISER"
  - "Building riser ..." → "FROM BUILDING RISER"
  - If the supply source text explicitly states the label, use it as-is (uppercased).

## Output JSON Schema

### Single-DB format (when file has only one board):
```json
{
  "incoming": {
    "kva": <float or null>,
    "phase": "<single_phase|three_phase or null>",
    "voltage": <230|400 or null>,
    "supply_source": "<sp_powergrid|landlord|null>",
    "incoming_label": "<supply source label for SLD diagram, e.g. 'FROM LANDLORD RISER', 'INCOMING FROM HDB ELECTRICAL RISER', or null>",
    "main_breaker": {
      "type": "<MCB|MCCB|ACB or null>",
      "rating_a": <int or null>,
      "poles": "<SPN|DP|TPN|4P or null>",
      "ka_rating": <int or null>,
      "characteristic": "<B|C|D or null>"
    },
    "cable": {
      "size_mm2": "<string or null>",
      "earth_mm2": "<string or null>",
      "type": "<PVC|XLPE|PVC/PVC|XLPE/PVC or null>",
      "cores": "<string e.g. '4 X 1 CORE' or null>",
      "description": "<full cable text or null>"
    },
    "elcb": {
      "type": "<ELCB|RCCB or null>",
      "rating_a": <int or null>,
      "poles": <2|4 or null>,
      "sensitivity_ma": <30|100|300 or null>
    },
    "busbar": {
      "rating_a": <int or null>,
      "type": "<COMB|COPPER or null>"
    },
    "metering": {
      "type": "<ct_meter|sp_meter or null>",
      "ct_ratio": "<string e.g. '100/5A' or null>",
      "protection_ct_ratio": "<string e.g. '100/5A' — ratio for PROTECTION CT, or null if not specified>",
      "protection_ct_class": "<string e.g. '5P10 20VA' — class for protection CT, or null>",
      "metering_ct_class": "<string e.g. 'CL1 5VA' — class for metering CT, or null>",
      "has_indicator_lights": "<bool — true if L1/L2/L3 indicator lights on potential fuse branch, default true for ct_meter>",
      "has_elr": "<bool — true if ELR (Earth Leakage Relay) present, default true for ct_meter>",
      "elr_spec": "<string e.g. '0-3A 0.2 SEC' — ELR specification, or null>",
      "has_ammeter": "<bool — true if ammeter with selector switch (ASS) present, default true for ct_meter>",
      "has_voltmeter": "<bool — true if voltmeter with selector switch (VSS) present, default true for ct_meter>",
      "voltmeter_range": "<string e.g. '0-500V' or null>",
      "ammeter_range": "<string e.g. '0-500A' or null>"
    },
    "outgoing_cable": {
      "size_mm2": "<string or null>",
      "description": "<full cable text from isolator to DB, if different from incoming cable, or null>"
    }
  },
  "outgoing_circuits": [
    {
      "id": "<circuit ID e.g. L1P1, S1, ISOL 1, Spare>",
      "description": "<load description>",
      "breaker": {
        "type": "<MCB|MCCB|ISOLATOR or null>",
        "rating_a": <int or null>,
        "poles": "<SPN|DP|TPN or null>",
        "ka_rating": <int or null>,
        "characteristic": "<B|C|D or null>"
      },
      "cable": "<full cable spec string or null>",
      "room": "<room/location or null>",
      "load_type": "<lighting|power|aircon|spare|isolator|heater|motor|other or null>",
      "phase": "<L1|L2|L3 or null>"
    }
  ],
  "client_info": {
    "name": "<client name or null>",
    "address": "<address or null>",
    "unit_number": "<unit number or null>",
    "drawing_no": "<drawing number or null>"
  }
}
```

### Multi-DB format (when file has multiple boards/sheets):
```json
{
  "incoming": {
    "kva": <float or null>,
    "phase": "<single_phase|three_phase or null>",
    "voltage": <230|400 or null>,
    "supply_source": "<sp_powergrid|landlord|null>",
    "incoming_label": "<supply source label for SLD diagram or null>",
    "main_breaker": { ... },
    "cable": { ... },
    "metering": { ... },
    "outgoing_cable": { ... }
  },
  "distribution_boards": [
    {
      "name": "<board name e.g. MSB, DB2, Lighting DB>",
      "breaker": {
        "type": "<MCB|MCCB or null>",
        "rating_a": <int or null>,
        "poles": "<SPN|DP|TPN|4P or null>",
        "ka_rating": <int or null>,
        "characteristic": "<B|C|D or null>"
      },
      "elcb": {
        "type": "<ELCB|RCCB or null>",
        "rating_a": <int or null>,
        "poles": <2|4 or null>,
        "sensitivity_ma": <30|100|300 or null>
      },
      "busbar": {
        "rating_a": <int or null>,
        "type": "<COMB|COPPER or null>"
      },
      "protection_groups": [
        {
          "phase": "<L1|L2|L3>",
          "rccb": {
            "type": "RCCB",
            "rating_a": <int>,
            "poles": <2|4>,
            "sensitivity_ma": <30|100|300>
          },
          "circuits": [
            {
              "id": "<circuit ID>",
              "description": "<load description>",
              "breaker": { ... },
              "cable": "<cable spec or null>",
              "load_type": "<load type>"
            }
          ]
        }
      ],
      "outgoing_circuits": [
        {
          "id": "<circuit ID>",
          "description": "<load description>",
          "breaker": { ... },
          "cable": "<cable spec or null>",
          "load_type": "<load type>",
          "phase": "<L1|L2|L3 or null>"
        }
      ]
    }
  ],
  "client_info": { ... }
}
```

## Rules
1. Extract ALL circuits — do not skip any rows including Spare and Isolator circuits.
2. Use `null` for any field that is uncertain or not present.
3. Normalize breaker types to uppercase: "MCB", "MCCB", "ACB", "ISOLATOR".
4. Normalize phase: "single_phase" or "three_phase".
5. Normalize poles: "SPN", "DP", "TPN", "4P" (uppercase).
6. For load_type, classify based on description keywords:
   - lighting/light/lamp/LED → "lighting"
   - socket/power/outlet/SSO/S/S/O → "power"
   - aircon/AC/air-con → "aircon"
   - spare → "spare"
   - isolator/ISOL → "isolator"
   - heater/water heater → "heater"
   - motor/pump/compressor → "motor"
   - everything else → "other"
7. If kVA is not stated but can be calculated from breaker rating × voltage, calculate it.
8. Preserve the original circuit ordering from the file.
9. Output ONLY valid JSON — no markdown, no explanation, no code fences.
10. **Multi-DB detection**: If the file has multiple sheets/tabs each representing a distribution board, OR has separate sections for different boards, use `distribution_boards` array. Otherwise use flat `outgoing_circuits`.
11. **Protection groups**: Group circuits into `protection_groups` when ANY of these conditions is met:
    a) The file has a "Protection Group" column (e.g., "RCCB L1", "RCCB L2", "RCCB L3").
    b) A three-phase DB has circuits assigned to specific phases (L1/L2/L3 or R/Y/B) AND has per-phase RCCBs.
    c) A three-phase DB has many circuits (≥9) with clear per-phase distribution — in Singapore practice, such boards typically use per-phase 2P RCCBs (40A 2P RCCB per phase with separate busbar).
    When creating protection_groups, each group has its own RCCB (typically 2P for single-phase groups within a 3-phase board) and its own busbar segment. Circuits NOT assigned to a specific phase go into the board's `outgoing_circuits`.
    **IMPORTANT — 4P RCCB rule**: If the schedule shows a SINGLE 4P RCCB (e.g., "63A 4P RCCB 30mA") for the entire board, do NOT split it into per-phase 2P RCCB groups. Instead, put ALL circuits into the board's `outgoing_circuits` (with their phase assignments) and set the board-level `elcb` to that 4P RCCB. Only create separate per-phase `protection_groups` when the schedule explicitly shows different/separate RCCBs per phase (e.g., "40A 2P RCCB" for each phase).
12. **Phase normalization**: Always normalize phase names to L1/L2/L3. Convert R→L1, Y→L2, B→L3, RED→L1, YELLOW→L2, BLUE→L3.
13. **Metering**: supply_source and metering are INDEPENDENT. For landlord supply with NO metering equipment mentioned → metering = null. If the schedule explicitly lists metering equipment (kWh meter, CT, SPPG, ammeter, voltmeter, ELR) → set metering accordingly (ct_meter or sp_meter). Do NOT change supply_source — a landlord supply can have CT metering.
14. **Outgoing cable**: If there are two different cables (e.g., one from riser to isolator, another from isolator to DB), capture the second cable in `outgoing_cable`. This is common in landlord supply installations."""


# ── File Type Detection ─────────────────────────────

_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
_CSV_EXTENSIONS = {".csv"}
_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".webp"}
_PDF_EXTENSIONS = {".pdf"}

_MIME_MAP = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".bmp": "image/bmp",
    ".tiff": "image/tiff",
    ".tif": "image/tiff",
    ".webp": "image/webp",
    ".pdf": "application/pdf",
}


def _get_file_extension(filename: str) -> str:
    """Extract lowercase file extension."""
    import os
    return os.path.splitext(filename)[1].lower()


def _detect_file_type(filename: str) -> str:
    """Detect file type from extension.

    Returns: "excel", "csv", "image", "pdf", or raises ValueError.
    """
    ext = _get_file_extension(filename)
    if ext in _EXCEL_EXTENSIONS:
        return "excel"
    if ext in _CSV_EXTENSIONS:
        return "csv"
    if ext in _IMAGE_EXTENSIONS:
        return "image"
    if ext in _PDF_EXTENSIONS:
        return "pdf"
    raise ValueError(f"Unsupported file format: {ext} (supported: .xlsx, .csv, .jpg, .png, .pdf)")


# ── Excel/CSV → Text Conversion ────────────────────

def _excel_to_text(file_bytes: bytes) -> str:
    """Convert Excel file bytes to a text representation of all sheets."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")

        rows_data: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cell_values = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if any(v.strip() for v in cell_values):
                rows_data.append(cell_values)

        if not rows_data:
            text_parts.append("(empty sheet)")
            continue

        # Format as tab-separated for clarity
        for row_data in rows_data:
            text_parts.append("\t".join(row_data))

    wb.close()
    return "\n".join(text_parts)


def _csv_to_text(file_bytes: bytes) -> str:
    """Convert CSV file bytes to text."""
    # Try UTF-8 first, then fallback to latin-1
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            # Validate it's parseable CSV
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows:
                return text
        except (UnicodeDecodeError, csv.Error):
            continue
    # Last resort: decode with replacement
    return file_bytes.decode("utf-8", errors="replace")


# ── Gemini API Call ─────────────────────────────────

async def _call_gemini_text(text_content: str, api_key: str | None = None) -> dict:
    """Call Gemini with text content for schedule extraction."""
    from google import genai
    from google.genai import types

    from app.config import settings

    resolved_key = api_key or settings.gemini_api_key
    client = genai.Client(api_key=resolved_key)

    logger.info("Calling Gemini for schedule extraction (text mode, %d chars)", len(text_content))

    response = await client.aio.models.generate_content(
        model=settings.gemini_model,
        contents=text_content,
        config=types.GenerateContentConfig(
            system_instruction=SCHEDULE_EXTRACTION_PROMPT,
            response_mime_type="application/json",
            temperature=0.1,
        ),
    )
    return json.loads(response.text)


async def _call_gemini_vision(
    file_bytes: bytes,
    mime_type: str,
    api_key: str | None = None,
) -> dict:
    """Call Gemini with image/PDF content for schedule extraction via Vision."""
    from google import genai
    from google.genai import types

    from app.config import settings

    resolved_key = api_key or settings.gemini_api_key
    client = genai.Client(api_key=resolved_key)

    logger.info("Calling Gemini for schedule extraction (vision mode, %s, %d bytes)", mime_type, len(file_bytes))

    # Multimodal content: text prompt + inline image/PDF
    b64_data = base64.b64encode(file_bytes).decode("utf-8")

    response = await client.aio.models.generate_content(
        model=settings.gemini_model,
        contents=[
            types.Content(
                parts=[
                    types.Part.from_text(
                        "Extract all electrical circuit schedule information from this file."
                    ),
                    types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                ],
            ),
        ],
        config=types.GenerateContentConfig(
            system_instruction=SCHEDULE_EXTRACTION_PROMPT,
            response_mime_type="application/json",
            temperature=0.1,
        ),
    )
    return json.loads(response.text)


# ── Post-processing: Auto-detect Protection Groups ──


def _auto_detect_protection_groups(data: dict) -> dict:
    """Post-process Gemini output: detect and create per-phase RCCB protection groups.

    For three-phase distribution boards where circuits are assigned to specific phases
    (L1/L2/L3) but Gemini did not create protection_groups, this function auto-groups
    them — matching Singapore LEW practice of per-phase 2P RCCB grouping.
    """
    dbs = data.get("distribution_boards")
    if not dbs:
        return data

    # Check if incoming is three-phase
    incoming = data.get("incoming", {})
    is_three_phase = incoming.get("phase") == "three_phase"
    if not is_three_phase:
        return data

    for db in dbs:
        # Skip if protection_groups already populated
        if db.get("protection_groups"):
            continue

        circuits = db.get("outgoing_circuits", [])
        if not circuits:
            continue

        # Count circuits per phase
        phase_map: dict[str, list[dict]] = {"L1": [], "L2": [], "L3": []}
        unassigned: list[dict] = []
        _PHASE_NORM = {
            "L1": "L1", "L2": "L2", "L3": "L3",
            "R": "L1", "Y": "L2", "B": "L3",
            "RED": "L1", "YELLOW": "L2", "BLUE": "L3",
        }

        for c in circuits:
            phase_raw = (c.get("phase") or "").upper().strip()
            phase = _PHASE_NORM.get(phase_raw)
            if phase:
                phase_map[phase].append(c)
            else:
                # Try to infer from circuit ID (e.g., L1P1, L2S1)
                cid = (c.get("id") or "").upper()
                inferred = None
                for prefix in ("L1", "L2", "L3"):
                    if cid.startswith(prefix):
                        inferred = prefix
                        break
                if inferred:
                    phase_map[inferred].append(c)
                else:
                    unassigned.append(c)

        # Only create protection groups if ≥2 phases have circuits and
        # total phase-assigned circuits ≥ 6 (typical 3-phase DB threshold)
        phases_with_circuits = sum(1 for v in phase_map.values() if v)
        total_assigned = sum(len(v) for v in phase_map.values())

        if phases_with_circuits >= 2 and total_assigned >= 6:
            # Determine per-phase RCCB rating from board's ELCB or default
            board_elcb = db.get("elcb", {}) or {}
            per_phase_rating = board_elcb.get("rating_a", 40)
            sensitivity = board_elcb.get("sensitivity_ma", 30)

            # Do NOT split into per-phase groups if the board has a 4P RCCB —
            # a 4P RCCB protects all phases together as a single device.
            board_elcb_poles = board_elcb.get("poles")
            if board_elcb_poles == 4 or board_elcb_poles == "4P":
                logger.info(
                    "Skipping auto protection-group split for DB '%s': board has 4P RCCB",
                    db.get("name", "?"),
                )
                continue

            protection_groups = []
            for phase in ("L1", "L2", "L3"):
                if not phase_map[phase]:
                    continue
                protection_groups.append({
                    "phase": phase,
                    "rccb": {
                        "type": "RCCB",
                        "rating_a": per_phase_rating,
                        "poles": 2,
                        "sensitivity_ma": sensitivity,
                    },
                    "circuits": phase_map[phase],
                })

            if protection_groups:
                db["protection_groups"] = protection_groups
                db["outgoing_circuits"] = unassigned
                # Remove board-level ELCB since it's now per-phase
                if db.get("elcb"):
                    del db["elcb"]
                logger.info(
                    "Auto-detected protection groups for DB '%s': %d groups, %d circuits",
                    db.get("name", "?"),
                    len(protection_groups),
                    total_assigned,
                )

    return data


def _merge_identical_4p_rccb_groups(data: dict) -> dict:
    """Safety net: merge protection_groups back when they all share the same 4P RCCB.

    If Gemini incorrectly split a single 4P RCCB into per-phase groups (each with
    the same RCCB spec and poles=4), merge them back into the board's outgoing_circuits
    and restore the board-level elcb.
    """
    dbs = data.get("distribution_boards")
    if not dbs:
        return data

    for db in dbs:
        pgs = db.get("protection_groups")
        if not pgs or len(pgs) < 2:
            continue

        # Check if ALL protection groups have the same RCCB spec with poles=4
        first_rccb = pgs[0].get("rccb", {})
        first_poles = first_rccb.get("poles")
        if first_poles not in (4, "4P"):
            continue

        all_same = all(
            pg.get("rccb", {}).get("type") == first_rccb.get("type")
            and pg.get("rccb", {}).get("rating_a") == first_rccb.get("rating_a")
            and pg.get("rccb", {}).get("poles") in (4, "4P")
            and pg.get("rccb", {}).get("sensitivity_ma") == first_rccb.get("sensitivity_ma")
            for pg in pgs
        )
        if not all_same:
            continue

        # Merge: collect all circuits from protection groups into outgoing_circuits
        merged_circuits: list[dict] = []
        for pg in pgs:
            phase = pg.get("phase")
            for c in pg.get("circuits", []):
                if phase and not c.get("phase"):
                    c["phase"] = phase
                merged_circuits.append(c)

        existing_circuits = db.get("outgoing_circuits", [])
        db["outgoing_circuits"] = merged_circuits + existing_circuits
        db["protection_groups"] = []

        # Restore board-level ELCB from the merged RCCB spec
        db["elcb"] = {
            "type": first_rccb.get("type", "RCCB"),
            "rating_a": first_rccb.get("rating_a"),
            "poles": 4,
            "sensitivity_ma": first_rccb.get("sensitivity_ma"),
        }

        logger.info(
            "Merged %d identical 4P RCCB protection_groups back into board-level ELCB for DB '%s' (%d circuits)",
            len(pgs),
            db.get("name", "?"),
            len(merged_circuits),
        )

    return data


# ── Public API ──────────────────────────────────────

async def extract_schedule_from_file(
    file_bytes: bytes,
    filename: str,
    api_key: str | None = None,
) -> dict[str, Any]:
    """Extract circuit schedule data from a file using Gemini AI.

    Supports Excel, CSV, Image, and PDF files.

    Args:
        file_bytes: Raw file content bytes.
        filename: Original filename (used for type detection).
        api_key: Optional Gemini API key override.

    Returns:
        {
            "success": bool,
            "file_type": "excel" | "csv" | "image" | "pdf",
            "raw_text": str,           # Text representation (for Excel/CSV) or "[image/pdf]"
            "extracted_data": dict,     # Gemini-extracted structured data
            "warnings": list[str],
        }
    """
    warnings: list[str] = []

    try:
        file_type = _detect_file_type(filename)
    except ValueError as e:
        return {
            "success": False,
            "file_type": "unknown",
            "raw_text": "",
            "extracted_data": {},
            "warnings": [str(e)],
            "error": str(e),
        }

    raw_text = ""
    extracted_data: dict = {}

    try:
        if file_type == "excel":
            raw_text = _excel_to_text(file_bytes)
            if not raw_text.strip() or raw_text.strip() == "(empty sheet)":
                return {
                    "success": False,
                    "file_type": file_type,
                    "raw_text": raw_text,
                    "extracted_data": {},
                    "warnings": ["Excel file appears to be empty"],
                    "error": "Empty Excel file",
                }
            extracted_data = await _call_gemini_text(raw_text, api_key)

        elif file_type == "csv":
            raw_text = _csv_to_text(file_bytes)
            if not raw_text.strip():
                return {
                    "success": False,
                    "file_type": file_type,
                    "raw_text": "",
                    "extracted_data": {},
                    "warnings": ["CSV file appears to be empty"],
                    "error": "Empty CSV file",
                }
            extracted_data = await _call_gemini_text(raw_text, api_key)

        elif file_type in ("image", "pdf"):
            raw_text = f"[{file_type}]"
            ext = _get_file_extension(filename)
            mime_type = _MIME_MAP.get(ext, "application/octet-stream")
            extracted_data = await _call_gemini_vision(file_bytes, mime_type, api_key)

    except json.JSONDecodeError as e:
        logger.error("Gemini returned invalid JSON for schedule extraction: %s", e)
        return {
            "success": False,
            "file_type": file_type,
            "raw_text": raw_text,
            "extracted_data": {},
            "warnings": [f"AI returned invalid JSON: {e}"],
            "error": f"JSON parse error: {e}",
        }
    except Exception as e:
        logger.error("Schedule extraction failed: %s", e, exc_info=True)
        return {
            "success": False,
            "file_type": file_type,
            "raw_text": raw_text,
            "extracted_data": {},
            "warnings": [f"Extraction failed: {e}"],
            "error": str(e),
        }

    # Post-process: merge incorrectly-split 4P RCCB groups (safety net — runs first)
    extracted_data = _merge_identical_4p_rccb_groups(extracted_data)
    # Post-process: auto-detect per-phase RCCB protection groups
    extracted_data = _auto_detect_protection_groups(extracted_data)

    # Validate basic structure (check both single-DB and multi-DB paths)
    has_circuits = bool(extracted_data.get("outgoing_circuits"))
    has_dbs = bool(extracted_data.get("distribution_boards"))
    if not has_circuits and not has_dbs:
        warnings.append("No outgoing circuits found in the file")
    if not extracted_data.get("incoming"):
        warnings.append("No incoming supply information found in the file")

    # Count total circuits across all paths
    circuit_count = len(extracted_data.get("outgoing_circuits", []))
    for _db in extracted_data.get("distribution_boards", []):
        circuit_count += len(_db.get("outgoing_circuits") or [])
        for _pg in (_db.get("protection_groups") or []):
            circuit_count += len(_pg.get("circuits") or [])
    logger.info(
        "Schedule extraction complete: file_type=%s, circuits=%d, warnings=%d",
        file_type, circuit_count, len(warnings),
    )

    return {
        "success": True,
        "file_type": file_type,
        "raw_text": raw_text,
        "extracted_data": extracted_data,
        "warnings": warnings,
    }


def format_extracted_schedule(result: dict) -> str:
    """Format extraction result as human-readable text for AI agent context.

    Args:
        result: Output from extract_schedule_from_file().

    Returns:
        Formatted text summary suitable for injection into agent system message.
    """
    if not result.get("success"):
        error = result.get("error", "Unknown error")
        return f"[File Analysis Failed] {error}"

    data = result.get("extracted_data", {})
    lines: list[str] = []
    lines.append(f"[Circuit Schedule File Analysis — {result.get('file_type', '?').upper()} file]")
    lines.append("")

    # Incoming supply
    incoming = data.get("incoming", {})
    if incoming:
        lines.append("## Incoming Supply")
        if incoming.get("kva"):
            lines.append(f"- Approved Load: {incoming['kva']} kVA")
        if incoming.get("phase"):
            phase_label = "Three-Phase 400V" if incoming["phase"] == "three_phase" else "Single-Phase 230V"
            lines.append(f"- Supply Type: {phase_label}")
        if incoming.get("supply_source"):
            lines.append(f"- Supply Source: {incoming['supply_source']}")

        mb = incoming.get("main_breaker", {})
        if mb:
            parts = [p for p in [
                mb.get("type"),
                f"{mb['rating_a']}A" if mb.get("rating_a") else None,
                mb.get("poles"),
                f"{mb['ka_rating']}kA" if mb.get("ka_rating") else None,
                f"Type {mb['characteristic']}" if mb.get("characteristic") else None,
            ] if p]
            if parts:
                lines.append(f"- Main Breaker: {' '.join(parts)}")

        elcb = incoming.get("elcb", {})
        if elcb:
            parts = [p for p in [
                elcb.get("type"),
                f"{elcb['rating_a']}A" if elcb.get("rating_a") else None,
                f"{elcb['sensitivity_ma']}mA" if elcb.get("sensitivity_ma") else None,
            ] if p]
            if parts:
                lines.append(f"- ELCB/RCCB: {' '.join(parts)}")

        cable = incoming.get("cable", {})
        if cable and cable.get("description"):
            lines.append(f"- Incoming Cable: {cable['description']}")

        busbar = incoming.get("busbar", {})
        if busbar and busbar.get("rating_a"):
            lines.append(f"- Busbar: {busbar['rating_a']}A {busbar.get('type', '')}")

        metering = incoming.get("metering", {})
        if metering and metering.get("type"):
            lines.append(f"- Metering: {metering['type']}")

    # Helper to format a circuit line
    def _fmt_circuit(oc: dict, idx: int) -> str:
        cid = oc.get("id", f"#{idx}")
        desc = oc.get("description", "Unknown")
        breaker = oc.get("breaker", {})
        br_parts = [p for p in [
            breaker.get("type"),
            f"{breaker['rating_a']}A" if breaker.get("rating_a") else None,
            breaker.get("poles"),
            f"Type {breaker['characteristic']}" if breaker.get("characteristic") else None,
        ] if p]
        br_str = " ".join(br_parts) if br_parts else "—"
        cable_str = oc.get("cable", "—") or "—"
        room_str = f" [{oc['room']}]" if oc.get("room") else ""
        return f"  {cid}: {desc}{room_str} | Breaker: {br_str} | Cable: {cable_str}"

    # Distribution boards (multi-DB)
    dbs = data.get("distribution_boards", [])
    if dbs:
        for db in dbs:
            db_name = db.get("name", "DB")
            pg_list = db.get("protection_groups", [])
            db_circuits = db.get("outgoing_circuits", [])
            total = len(db_circuits) + sum(len(pg.get("circuits", [])) for pg in pg_list)
            lines.append(f"\n## {db_name} ({total} circuits)")

            if pg_list:
                for pg in pg_list:
                    phase = pg.get("phase", "?")
                    rccb = pg.get("rccb", {})
                    rccb_str = f"{rccb.get('rating_a', '?')}A {rccb.get('poles', '?')}P RCCB ({rccb.get('sensitivity_ma', '?')}mA)"
                    pg_circuits = pg.get("circuits", [])
                    lines.append(f"  ### Protection Group {phase} — {rccb_str} ({len(pg_circuits)} circuits)")
                    for j, c in enumerate(pg_circuits, 1):
                        lines.append(f"  {_fmt_circuit(c, j)}")

            for j, oc in enumerate(db_circuits, 1):
                lines.append(_fmt_circuit(oc, j))

    # Outgoing circuits (single-DB)
    circuits = data.get("outgoing_circuits", [])
    if circuits:
        lines.append(f"\n## Circuit Schedule ({len(circuits)} circuits)")
        for i, oc in enumerate(circuits, 1):
            lines.append(_fmt_circuit(oc, i))

    # Client info
    client = data.get("client_info", {})
    if client and any(client.values()):
        lines.append("\n## Client Information")
        if client.get("name"):
            lines.append(f"- Client: {client['name']}")
        if client.get("address"):
            lines.append(f"- Address: {client['address']}")
        if client.get("unit_number"):
            lines.append(f"- Unit: {client['unit_number']}")

    # Warnings
    if result.get("warnings"):
        lines.append(f"\n## Warnings ({len(result['warnings'])})")
        for w in result["warnings"]:
            lines.append(f"  - {w}")

    return "\n".join(lines)
