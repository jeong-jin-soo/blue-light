"""Tests for circuit schedule file parser (schedule_parser.py).

Tests the file type detection, text conversion, and formatting functions.
Gemini API calls are mocked to avoid external dependencies.
"""

import asyncio
import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.sld.schedule_parser import (
    _csv_to_text,
    _detect_file_type,
    _excel_to_text,
    extract_schedule_from_file,
    format_extracted_schedule,
)


def _run(coro):
    """Helper to run async functions in sync tests."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


# ── File Type Detection ────────────────────────────────


class TestDetectFileType:
    def test_xlsx(self):
        assert _detect_file_type("schedule.xlsx") == "excel"

    def test_xls(self):
        assert _detect_file_type("old_file.xls") == "excel"

    def test_csv(self):
        assert _detect_file_type("circuits.csv") == "csv"

    def test_jpg(self):
        assert _detect_file_type("photo.jpg") == "image"

    def test_jpeg(self):
        assert _detect_file_type("photo.JPEG") == "image"

    def test_png(self):
        assert _detect_file_type("screenshot.png") == "image"

    def test_pdf(self):
        assert _detect_file_type("document.pdf") == "pdf"

    def test_unsupported(self):
        with pytest.raises(ValueError, match="Unsupported file format"):
            _detect_file_type("file.docx")

    def test_case_insensitive(self):
        assert _detect_file_type("SCHEDULE.XLSX") == "excel"
        assert _detect_file_type("Photo.PNG") == "image"


# ── CSV to Text ───────────────────────────────────────


class TestCsvToText:
    def test_basic_csv(self):
        csv_bytes = b"Circuit,Load,Breaker\nL1P1,Lighting,10A MCB\n"
        result = _csv_to_text(csv_bytes)
        assert "Circuit" in result
        assert "L1P1" in result
        assert "Lighting" in result

    def test_utf8_bom(self):
        csv_bytes = b"\xef\xbb\xbfCircuit,Load\nS1,Test\n"
        result = _csv_to_text(csv_bytes)
        assert "Circuit" in result
        assert "S1" in result

    def test_empty_csv(self):
        result = _csv_to_text(b"")
        assert result == ""


# ── Excel to Text ─────────────────────────────────────


class TestExcelToText:
    def test_basic_excel(self):
        """Test with a real openpyxl workbook."""
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Main Circuit"
        ws.append(["Supply Source", "SP PowerGrid"])
        ws.append(["Main Breaker", "63A TPN MCB 10kA Type B"])

        ws2 = wb.create_sheet("Circuit Schedule")
        ws2.append(["Circuit", "Load", "Breaker", "Cable"])
        ws2.append(["L1P1", "8 Nos Lighting Points", "10A MCB Type B", "2x1.5mm PVC"])
        ws2.append(["Spare", "Spare", "", ""])

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        result = _excel_to_text(excel_bytes)
        assert "Main Circuit" in result
        assert "SP PowerGrid" in result
        assert "Circuit Schedule" in result
        assert "L1P1" in result
        assert "8 Nos Lighting Points" in result
        assert "Spare" in result


# ── Format Extracted Schedule ──────────────────────────


class TestFormatExtractedSchedule:
    def test_success_format(self):
        result = {
            "success": True,
            "file_type": "excel",
            "extracted_data": {
                "incoming": {
                    "kva": 69.282,
                    "phase": "three_phase",
                    "main_breaker": {
                        "type": "MCB",
                        "rating_a": 63,
                        "poles": "TPN",
                        "ka_rating": 10,
                        "characteristic": "B",
                    },
                    "elcb": {
                        "type": "ELCB",
                        "rating_a": 63,
                        "sensitivity_ma": 30,
                    },
                    "busbar": {"rating_a": 100, "type": "COMB"},
                },
                "outgoing_circuits": [
                    {
                        "id": "L1P1",
                        "description": "8 Nos Lighting Points",
                        "breaker": {"type": "MCB", "rating_a": 10, "characteristic": "B"},
                        "cable": "2x1C 1.5mm PVC",
                    },
                    {
                        "id": "Spare",
                        "description": "Spare",
                        "breaker": {},
                        "cable": None,
                    },
                ],
            },
            "warnings": [],
        }

        text = format_extracted_schedule(result)
        assert "EXCEL file" in text
        assert "69.282 kVA" in text
        assert "Three-Phase 400V" in text
        assert "MCB 63A TPN 10kA Type B" in text
        assert "ELCB 63A 30mA" in text
        assert "100A COMB" in text
        assert "L1P1" in text
        assert "8 Nos Lighting Points" in text
        assert "2 circuits" in text

    def test_failure_format(self):
        result = {
            "success": False,
            "error": "Unsupported file format: .docx",
        }
        text = format_extracted_schedule(result)
        assert "File Analysis Failed" in text
        assert ".docx" in text

    def test_with_warnings(self):
        result = {
            "success": True,
            "file_type": "csv",
            "extracted_data": {
                "outgoing_circuits": [
                    {"id": "S1", "description": "Lighting", "breaker": {}, "cable": None},
                ],
            },
            "warnings": ["No incoming supply information found in the file"],
        }
        text = format_extracted_schedule(result)
        assert "Warnings" in text
        assert "No incoming supply information" in text


# ── Extract Schedule from File (Mocked) ──────────────


class TestExtractScheduleFromFile:
    def test_unsupported_format(self):
        result = _run(extract_schedule_from_file(b"data", "file.docx"))
        assert result["success"] is False
        assert "Unsupported" in result["error"]

    def test_empty_csv(self):
        result = _run(extract_schedule_from_file(b"", "empty.csv"))
        assert result["success"] is False
        assert "empty" in result["error"].lower()

    def test_csv_extraction_mocked(self):
        """Test CSV extraction with mocked Gemini call."""
        mock_response = {
            "incoming": {"kva": 40, "phase": "single_phase"},
            "outgoing_circuits": [
                {"id": "S1", "description": "Lighting", "breaker": {"type": "MCB", "rating_a": 10}},
            ],
        }

        with patch("app.sld.schedule_parser._call_gemini_text", new_callable=AsyncMock) as mock_gemini:
            mock_gemini.return_value = mock_response
            result = _run(extract_schedule_from_file(
                b"Circuit,Load\nS1,Lighting\n",
                "test.csv",
            ))
            assert result["success"] is True
            assert result["file_type"] == "csv"
            assert len(result["extracted_data"]["outgoing_circuits"]) == 1
            mock_gemini.assert_called_once()

    def test_image_extraction_mocked(self):
        """Test image extraction with mocked Gemini Vision call."""
        mock_response = {
            "incoming": {"kva": 69, "phase": "three_phase"},
            "outgoing_circuits": [
                {"id": "L1P1", "description": "Socket", "breaker": {"type": "MCB", "rating_a": 20}},
            ],
        }

        with patch("app.sld.schedule_parser._call_gemini_vision", new_callable=AsyncMock) as mock_gemini:
            mock_gemini.return_value = mock_response
            result = _run(extract_schedule_from_file(
                b"\x89PNG\r\n\x1a\n",  # PNG header bytes
                "schedule.png",
            ))
            assert result["success"] is True
            assert result["file_type"] == "image"
            assert result["raw_text"] == "[image]"
            mock_gemini.assert_called_once()

    def test_excel_extraction_mocked(self):
        """Test Excel extraction with mocked Gemini call."""
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Circuit", "Load"])
        ws.append(["S1", "Lighting"])
        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        mock_response = {
            "incoming": {"kva": 20, "phase": "single_phase"},
            "outgoing_circuits": [
                {"id": "S1", "description": "Lighting", "breaker": {"type": "MCB", "rating_a": 10}},
            ],
        }

        with patch("app.sld.schedule_parser._call_gemini_text", new_callable=AsyncMock) as mock_gemini:
            mock_gemini.return_value = mock_response
            result = _run(extract_schedule_from_file(excel_bytes, "test.xlsx"))
            assert result["success"] is True
            assert result["file_type"] == "excel"
            assert "S1" in result["raw_text"]
            mock_gemini.assert_called_once()

    def test_gemini_json_error(self):
        """Test handling of Gemini returning invalid JSON."""
        with patch("app.sld.schedule_parser._call_gemini_text", new_callable=AsyncMock) as mock_gemini:
            mock_gemini.side_effect = json.JSONDecodeError("Expecting value", "", 0)
            result = _run(extract_schedule_from_file(
                b"Circuit,Load\nS1,Lighting\n",
                "test.csv",
            ))
            assert result["success"] is False
            assert "JSON" in result["error"]
